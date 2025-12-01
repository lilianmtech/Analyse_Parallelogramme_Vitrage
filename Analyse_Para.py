import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl as xl
from matplotlib.patches import Polygon as MplPolygon
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.image as mpimg
import sympy as sy
from math import *
from numpy import *
from shapely import *
from scipy.optimize import minimize_scalar
import builtins
import io
import requests
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

class Vitrage:
    def __init__(self,cadre_0,cadre_def,Gamme,raico,pf,calage_lateral='Sans'):
        self.Gamme = Gamme
        self.A = cadre_def['A']
        self.B = cadre_def['B']
        self.C = cadre_def['C']
        self.D = cadre_def['D']
        self.A0 = cadre_0['A']
        self.B0 = cadre_0['B']
        self.C0 = cadre_0['C']
        self.D0 = cadre_0['D']

        self.calage_lateral = calage_lateral
        self.p = pf
        self.raico = raico

        self.cale_bas = 100
        self.cale_laterale = 100

    def DiffHorsPlan(self):
        "Vecteur normal"
        a = sy.Symbol('a')
        b = 1
        c = sy.Symbol('c')
        
        "Coeffiecient directeur de AB"
        x1 = self.B[0]-self.A[0]
        y1 = self.B[1]-self.A[1]
        z1 = self.B[2]-self.A[2]
        
        "Coeffiecient directeur de AD"
        x2 = self.D[0]-self.A[0]
        y2 = self.D[1]-self.A[1]
        z2 = self.D[2]-self.A[2]
        
        "Détermination des coef du vecteur normal"
        eq1 = sy.Eq(a*x1 + b*y1 + c*z1 , 0)
        eq2 = sy.Eq(a*x2 + b*y2 + c*z2 , 0)
        sol1 = sy.solve((eq1,eq2),(a,c))
        
        if sol1 == []:
            distance = 'None'
        else:
            a=float(sol1[a])
            c=float(sol1[c])
            
            "Détermination des coef de l'équation de plan"
            d = -a*self.A[0]-b*self.A[1]-c*self.A[2]
            
            "Distance entre point et plan"
            distance = abs(a*self.C[0]+b*self.C[1]+c*self.C[2]+d)/sqrt(a**2+b**2+c**2)
            
        #-----------Distance entre le sommet et le centre de la diagonale---------
        X = sy.Symbol('X')
        
        O = []
        
        "Coefficient droite AC"
        a1 = (self.C[1]-self.A[1])/(self.C[0]-self.A[0])
        b1 = self.C[1] - (a1*self.C[0])
        
        "Coefficient droite BD"
        a2 = (self.D[1]-self.B[1])/(self.D[0]-self.B[0])
        b2 = self.D[1] - (a2*self.D[0])
        
        eq3 = sy.Eq(a1*X + b1 - a2*X - b2 , 0)
        sol3 = sy.solve(eq3,X)
        O.append(float(sol3[0]))
        O.append(a1*O[0]+b1)
        
        LO = sqrt((O[0]-self.C[0])**2+(O[1]-self.C[1])**2)
        
        if abs(distance)<abs(LO/75):
            Gauchissement = 'OK'
        else:
            Gauchissement = 'NOK'

        return distance
    
    def inner_rect_with_offsets(self,quad, offset_top, offset_bottom, offset_lr):
        """
        quad : Polygon à 4 sommets
        offset_top : marge intérieure en haut
        offset_bottom : marge intérieure en bas
        offset_lr : marge intérieure gauche/droite (même valeur)
        """

        # Enveloppe du quadrilatère
        minx, miny, maxx, maxy = quad.bounds
        width  = maxx - minx
        height = maxy - miny

        # Centre du quadrilatère
        cx = (minx + maxx) / 2
        cy = (miny + maxy) / 2

        # Facteurs d'échelle nécessaires
        sx = (width - 2 * offset_lr) / width
        sy = (height - offset_top - offset_bottom) / height

        # Étape 1 : réduction (scale)
        inner = affinity.scale(quad, xfact=sx, yfact=sy, origin=(cx, cy))

        # Étape 2 : correction verticale
        translate_y = (offset_bottom - offset_top) / 2
        inner = affinity.translate(inner, yoff=translate_y)

        return inner


    def extend_line(self,line, factor=1000):
        """
        Étend une ligne dans les deux directions.
        factor: facteur d'extension (1000 = très long)
        """
        # Obtenir les coordonnées du segment
        coords = list(line.coords)
        start, end = coords[0], coords[-1]
    
        # Calculer le vecteur directionnel
        dx = end[0] - start[0]
        dy = end[1] - start[1]
    
        # Étendre le segment dans les deux directions
        new_start = (start[0] - dx * factor, start[1] - dy * factor)
        new_end = (end[0] + dx * factor, end[1] + dy * factor)
    
        return LineString([new_start, new_end])

    def find_intersection_point(self,segment1, segment2, extend_if_needed=True):
        """
        Trouve le point d'intersection entre deux segments.
        Si pas d'intersection, étend les segments pour trouver l'intersection.
        """
        # Vérifier si les segments se croisent déjà
        if segment1.intersects(segment2):
            intersection = segment1.intersection(segment2)
            if isinstance(intersection, Point):
                return intersection
    
        # Si pas d'intersection et extension autorisée
        if extend_if_needed:
            # Étendre les deux segments
            extended1 = self.extend_line(segment1)
            extended2 = self.extend_line(segment2)
        
            # Vérifier l'intersection des segments étendus
            if extended1.intersects(extended2):
                intersection = extended1.intersection(extended2)
                if isinstance(intersection, Point):
                    return intersection
    
        return None, False

    def inner_quad_with_offsets(self,cadre, offset_top=0.0, offset_bottom=0.0, offset_lr=0.0):
        A0,B0,C0,D0,ini =list(cadre.exterior.coords)
        A_B=offset_curve(LineString([A0,B0]), -offset_top)
        D_A =offset_curve(LineString([D0,A0]), -offset_lr)
        C_B=offset_curve(LineString([C0,B0]), offset_lr)
        D_C=offset_curve(LineString([D0,C0]), offset_bottom)
    
        A = self.find_intersection_point(A_B,D_A)
        B = self.find_intersection_point(A_B,C_B)
        C = self.find_intersection_point(C_B,D_C)
        D = self.find_intersection_point(D_C,D_A)
    
        return Polygon([A,B,C,D])

    def distance_angle(self,angle, vitrage, cote, origine, support):
        rotated = affinity.rotate(vitrage, angle, origin=origine)
        A_V, B_V, C_V, D_V, ini = list(rotated.exterior.coords)
        if cote == 'gauche' :
            line_vit = LineString([A_V, D_V])
        elif cote == 'droite' :
            line_vit = LineString([B_V, C_V])
        return line_vit.distance(support)

    def Dechaussement(self):
        
        #--------------------Defintion des bornes--------------------
        
        'Mise en place d''un dictionnaire pour les paramétres'
        l = ('Pos_vit_lat','Pos_vit_haut','Tolerance epine','Tolerance traverse','Calage lateral')
        Raico={}
        
        if self.p > 7.0 :
            pf = 12.0
        elif  5.0 < self.p <= 7.0 :
            pf = 9.0
        elif self.p <= 5.0 :
            pf = 6.0
        
        if self.calage_lateral == 'Sans':
            Tm1 = self.raico[0]
            Tm2 = self.raico[1]
            Tt1 = self.raico[2]
            Tt2 = self.raico[3]

            pos_vit_lat = (self.Gamme/2-(Tm1+pf+Tm2))/2+Tm2
            pos_vit_haut = (self.Gamme/2-(Tt1+pf+Tt2))/2+Tt2
            jeu_borne_lat = (self.Gamme/2-(Tm1+pf+Tm2))/2
            jeu_borne_haut = (self.Gamme/2-(Tt1+pf+Tt2))/2

            donnees = (pos_vit_lat, pos_vit_haut, jeu_borne_lat, jeu_borne_haut, 0)

        elif self.calage_lateral == 'Avec':
            Tm = self.raico[0]
            Ca = self.raico[1]
            Jc = self.raico[2]
            Tt1 = self.raico[3]
            Tt2 = self.raico[4]
            Jv = self.raico[5]
        
            pos_vit_lat = Ca+Jc
            pos_vit_haut = self.Gamme/2-(Tt2+pf+Tt1+Jv)+Tt2
            jeu_borne_lat = self.Gamme/2 - (Ca+Jc+pf+Tm)
            jeu_borne_haut = self.Gamme/2-(Tt1+pf+Tt2+Jv)

            donnees = (pos_vit_lat, pos_vit_haut, jeu_borne_lat, jeu_borne_haut, Ca)

        for i in range(0,len(l)):
            Raico[l[i]] = donnees[i]
        
        #-----------Definition des cales---------

        longueur_traverse_basse = sqrt((self.D[0] - self.C[0])**2 + (self.D[1] - self.C[1])**2)

        gauche_0 = LineString([(self.D0[0],self.D0[1]),(self.C0[0],self.C0[1])]).interpolate(self.cale_bas)
        gauche = LineString([(self.D[0],self.D[1]),(self.C[0],self.C[1])]).interpolate(self.cale_bas)
        droite_0 = LineString([(self.C0[0],self.C0[1]),(self.D0[0],self.D0[1])]).interpolate(self.cale_bas)
        droite = LineString([(self.C[0],self.C[1]),(self.D[0],self.D[1])]).interpolate(self.cale_bas)
        T_gauche = gauche_0.y-gauche.y
        T_droite = droite_0.y-droite.y
        
        traverse_basse_gauche = LineString([(self.D[0],self.D[1]),(self.C[0],self.C[1])])
        TG = traverse_basse_gauche.parallel_offset(13.0,'left',join_style=2)
        TGS = traverse_basse_gauche.interpolate(self.cale_bas)
        Distance_cale_gauche=TG.project(TGS)
        Support_cale_gauche=TG.interpolate(Distance_cale_gauche)

        traverse_basse_droite = LineString([(self.C[0],self.C[1]),(self.D[0],self.D[1])])
        TD = traverse_basse_droite.parallel_offset(13.0,'right',join_style=2)
        TDS = traverse_basse_droite.interpolate(self.cale_bas)
        Distance_cale_droite=TD.project(TDS)
        Support_cale_droite=TD.interpolate(Distance_cale_droite)

        support_bas = {
    "T_gauche": T_gauche,
    "gauche":   Support_cale_gauche,
    "T_droite": T_droite,
    "droite":   Support_cale_droite
}

        cale_lateral_gauche = LineString([(self.A[0],self.A[1]),(self.D[0],self.D[1])])
        BG = cale_lateral_gauche.parallel_offset(Raico['Calage lateral'],'left',join_style=2)
        BGS = cale_lateral_gauche.interpolate(self.cale_laterale)
        Distance_support_gauche=BG.project(BGS)
        Support_lateral_gauche=BG.interpolate(Distance_support_gauche)

        cale_lateral_droite = LineString([(self.B[0],self.B[1]),(self.C[0],self.C[1])])
        BD = cale_lateral_droite.parallel_offset(Raico['Calage lateral'],'right',join_style=2)
        BDS = cale_lateral_droite.interpolate(self.cale_laterale)
        Distance_support_droite=BD.project(BDS)
        Support_lateral_droite=BD.interpolate(Distance_support_droite)

        support_laterale = {
            "gauche":   Support_lateral_gauche,
            "droite":   Support_lateral_droite
        }


        #-----------Definition du vitrage---------

        cadre_0 = Polygon([
        list(self.A0[0:2]),
        list(self.B0[0:2]),
        list(self.C0[0:2]),
        list(self.D0[0:2])
    ])  

        Vitrage = self.inner_quad_with_offsets( cadre_0, offset_top=Raico['Pos_vit_haut'], offset_bottom=13.0,offset_lr=Raico['Pos_vit_lat'])
              
        
        #-----------Mise en mouvement---------           
        if self.C0[1] >= self.C[1]:
            #longueur_traverse_basse=hypoténuse
            cote_support_bas = 'gauche'
            signe_rotation = -1
            #Angle positif → rotation horaire (vers la droite)
        else :
            cote_support_bas = 'droite'
            signe_rotation = 1
            #Angle négatif → rotation anti‑horaire (vers la gauche)
        
        Angle=degrees(asin((abs(self.D[1]-self.C[1]))/longueur_traverse_basse))

        #Translation du vitrage
        Vitrage_T = affinity.translate(Vitrage,xoff=0.0, yoff=-support_bas['T_'+cote_support_bas], zoff=0.0)
        #Rotation du vitrage'
        Vitrage_T_R = affinity.rotate(Vitrage_T, signe_rotation*Angle, origin=(support_bas[cote_support_bas]))

        contact_cale = 'Non'

        if self.calage_lateral == 'Avec' :
            #On vérifie si conflit avec les cales :
            if Vitrage_T_R.distance(support_laterale['droite'])==0:
                func = lambda a: self.distance_angle(a, Vitrage_T, 'droite',support_bas[cote_support_bas], support_laterale['droite'])
                res = minimize_scalar(func, bounds=(-10, 10), method='bounded')
                Vitrage_T_R = affinity.rotate(Vitrage_T, res.x, origin=support_bas[cote_support_bas])
                contact_cale = 'droite'

            elif Vitrage_T_R.distance(support_laterale['gauche'])==0:
                func = lambda a: self.distance_angle(a, Vitrage_T, 'gauche',support_bas[cote_support_bas], support_laterale['gauche'])
                res = minimize_scalar(func, bounds=(-10, 10), method='bounded')
                Vitrage_T_R = affinity.rotate(Vitrage_T, res.x, origin=support_bas[cote_support_bas])
                contact_cale = 'gauche'
    
        #-----------Vérifications des limites---------     
        
        A_V, B_V, C_V, D_V, ini = list(Vitrage_T_R.exterior.coords)

        Cadre = Polygon([
            list(self.A[0:2]),
            list(self.B[0:2]),
            list(self.C[0:2]),
            list(self.D[0:2])
        ])

        if self.calage_lateral == 'Avec':
            Borne_ext = self.inner_quad_with_offsets( Cadre, offset_top=Raico['Pos_vit_haut']-Raico['Tolerance traverse'], offset_bottom=13.0,offset_lr=Raico['Calage lateral'])
            Borne_int = self.inner_quad_with_offsets( Cadre, offset_top=Raico['Pos_vit_haut']+Jv, offset_bottom=13.0,offset_lr=Raico['Pos_vit_lat']+Raico['Tolerance epine'])
        else :
            Borne_ext = self.inner_quad_with_offsets( Cadre, offset_top=Raico['Pos_vit_haut']-Raico['Tolerance traverse'], offset_bottom=13.0,offset_lr=Raico['Pos_vit_lat']-Raico['Tolerance epine'])
            Borne_int = self.inner_quad_with_offsets( Cadre, offset_top=Raico['Pos_vit_haut']+Raico['Tolerance traverse'], offset_bottom=13.0,offset_lr=Raico['Pos_vit_lat']+Raico['Tolerance epine'])


        A_BE, B_BE, C_BE, D_BE, ini = list(Borne_ext.exterior.coords)
        A_BI, B_BI, C_BI, D_BI, ini = list(Borne_int.exterior.coords)

        Ctrl_gauche_BE = A_V[0] - A_BE[0]
        Ctrl_gauche_BI = A_BI[0] - A_V[0]
        Ctrl_droite_BE = B_BE[0] - B_V [0]
        Ctrl_droite_BI =  B_V[0] - B_BI[0]
        Ctrl_haut_BE = builtins.min(A_BE[1] - A_V [1],B_BE[1] - B_V [1])
        Ctrl_haut_BI = builtins.min( A_V [1] - A_BI[1],B_V [1] - B_BI[1])
        
        if self.calage_lateral == 'Avec' :
            if contact_cale == 'droite':
                Ctrl_droite_BE = 0

            elif contact_cale == 'gauche':
                Ctrl_gauche_BE = 0

        Vitrage_T_R = Polygon(Vitrage_T_R)
        
        data = [round(builtins.min(Ctrl_gauche_BE,Ctrl_gauche_BI),1),round(builtins.min(Ctrl_droite_BE,Ctrl_droite_BI),1),round(builtins.min(Ctrl_haut_BE,Ctrl_haut_BI),1)]
        graph = [Cadre, Borne_ext, Vitrage_T_R, Borne_int]
        
        return data, graph

#-----------------------Importation des points d'origine----------------------"
def CoordPoint(doc,feuille):
    document =xl.load_workbook(doc,data_only=True)
    sheet = document[feuille]
    max_row = sheet.max_row
    Coord = {}
    Depl = {}
    labels = []
    for row in range(4, max_row+1):
        label = sheet.cell(row, column=3).value
        labels.append(label)
        
    #Attribuer les données
    i = 0
    for row in range(4, max_row+1) :
        X_Origin = sheet.cell(row, column=4).value
        Y_Origin = sheet.cell(row, column=5).value
        Z_Origin = sheet.cell(row, column=6).value
        coord_origin = [X_Origin,Y_Origin,Z_Origin]
        Coord[labels[i]] = coord_origin
        
        X_Depl = sheet.cell(row, column=7).value 
        Y_Depl = sheet.cell(row, column=8).value 
        Z_Depl = sheet.cell(row, column=9).value
        deplacement = [X_Depl,Y_Depl,Z_Depl]
        Depl[labels[i]] = deplacement
        i = 1+i
        
    return Coord, Depl

#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"

#-------------------------------changement de repére--------------------------

def rot_repere(angle,point):
    y = cos(radians(-angle))*point[1]-sin(radians(-angle))*point[2]
    z = sin(radians(-angle))*point[1]+cos(radians(-angle))*point[2]
    return[point[0], y, z]

#---------------------------Chargement des vitrages---------------------------"
def Dechaussement_vitrage(uploaded_file, option_calage, Raico):
    Feuille_coord = 'Coord_Points'
    Coord,Depl = CoordPoint(uploaded_file,Feuille_coord)
    
    document =xl.load_workbook(uploaded_file,data_only=True)
    sheet = document['Cadres']
    max_row = sheet.max_row
    datas = {
        'ID vitrage': [],
        'Gamme': [],
        'Demi-périmètre (m)': [],
        'Ecart minimal // bornes gauche (mm)': [],
        'Ecart minimal // bornes droite (mm)': [],
        'Ecart minimal // bornes hautes (mm)': []
    }

    graphs = {
        'ID vitrage': [],
        'Graph': [],
    }

    for row in range(3, max_row+1) :
        if sheet.cell(row+1, column=2).value == None:
            break
        
        labels = {
            "A": sheet.cell(row + 1, column=2).value,
            "B": sheet.cell(row + 1, column=3).value,
            "C": sheet.cell(row + 1, column=4).value,
            "D": sheet.cell(row + 1, column=5).value,
        }
        Gamme = sheet.cell(row + 1, column=6).value        
        
        cadre_def = {}
        cadre_0 = {}
        
        for key, label in labels.items():
            print([label][0],[labels["D"]][0])
            x = Coord[label][0] + Depl[label][0] - (Coord[labels["D"]][0] + Depl[labels["D"]][0])
            if key == "D":  # cas particulier pour D
                x = 0

            if key == "A":
                y = Coord[label][1] - Depl[labels["D"]][1] - (Coord[labels["D"]][1] - Depl[labels["D"]][1])
            elif key == "B":
                y = Coord[label][1] - Depl[labels["C"]][1] - (Coord[labels["D"]][1] - Depl[labels["D"]][1])
            elif key == "C":
                y = Coord[label][1] - Depl[labels["C"]][1] - (Coord[labels["D"]][1] - Depl[labels["D"]][1])
            else:  # D
                y = Coord[label][1] - Depl[label][1] - (Coord[labels["D"]][1] - Depl[labels["D"]][1])

            z = Coord[label][2] + Depl[label][2]

            cadre_def[key] = [x, y, z]

            x_0 = Coord[label][0] - Coord[labels["D"]][0]
            y_0 = Coord[label][1] - Coord[labels["D"]][1]
            z_0 = Coord[label][2]

            cadre_0[key] = [x_0, y_0, z_0]
        
        angle = degrees(arctan(cadre_0['A'][2]/cadre_0['A'][1]))
        print(angle)

        for key, coords in cadre_def.items():
             cadre_def[key] = rot_repere(angle,coords)

        for key, coords in cadre_0.items():
             cadre_0[key] = rot_repere(angle,coords)

        H = abs(cadre_0['A'][1]-cadre_0['D'][1])
        L = abs(cadre_0['A'][0]-cadre_0['B'][0])
        diag = sqrt(H**2+L**2)
        pf = L/1000 + H/1000
        V = Vitrage(cadre_0,cadre_def,Gamme,Raico,pf,calage_lateral=option_calage,)
        data, graph = V.Dechaussement()

        datas['ID vitrage'].append(str(labels["A"])+' / '+str(labels["B"])+' / '+str(labels["C"])+' / '+str(labels["D"]))
        datas['Gamme'].append(Gamme)
        datas['Demi-périmètre (m)'].append(round(pf,1))
        datas['Ecart minimal // bornes gauche (mm)'].append(data[0])
        datas['Ecart minimal // bornes droite (mm)'].append(data[1])
        datas['Ecart minimal // bornes hautes (mm)'].append(data[2])

        graphs['ID vitrage'].append(str(labels["A"])+' / '+str(labels["B"])+' / '+str(labels["C"])+' / '+str(labels["D"]))
        graphs['Graph'].append(graph)
        
    return datas, graphs

#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

#-----------------------visualiser les quadrilatères avec zoom----------------

def visualiser_quadrilateres(quadrilateres, titre_page=None):
    """Crée deux graphiques avec zoom adaptatif sur les coins supérieurs"""
    # Format A4 en paysage : 11.69 x 8.27 pouces
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11.69, 8.27))
    
    # Titre de la page si fourni
    if titre_page:
        fig.suptitle(titre_page, fontsize=16, fontweight='bold', y=0.98)
    
    couleurs =["#000000","#FF0000", '#006D8F', "#FF0000"]
    titre_graph =['Cadre','Limite','Vitrage','Limite']
    
    # Tracer les quadrilatères sur les deux axes - contours uniquement
    for ax in [ax1, ax2]:
        for i, quad in enumerate(quadrilateres):
            x, y = quad.exterior.xy
            patch = MplPolygon(list(zip(x, y)), facecolor='none', 
                             edgecolor=couleurs[i], linewidth=1.0,
                             label=titre_graph[i])
            ax.add_patch(patch)
        
        ax.set_aspect('equal')
        ax.grid(True, alpha=0.3)
        ax.legend()
    
    # Zoom adaptatif sur le coin supérieur gauche
    quad1 = quadrilateres[2]
    x1, y1 = quad1.exterior.xy
    
    # Trouver le coin supérieur gauche (point avec x min et y max)
    points1 = list(zip(x1, y1))
    coin_haut_gauche = builtins.min(points1, key=lambda p: p[0] - p[1])  # x minimum, y maximum
    
    # Calculer la taille de la zone de zoom basée sur la taille du quadrilatère
    largeur_quad1 = builtins.max(x1) - builtins.min(x1)
    hauteur_quad1 = builtins.max(y1) - builtins.min(y1)
    marge = builtins.max(largeur_quad1, hauteur_quad1) * 0.01  # 1% de marge
    
    ax1.set_xlim(coin_haut_gauche[0] - marge, coin_haut_gauche[0] + marge)
    ax1.set_ylim(coin_haut_gauche[1] - marge, coin_haut_gauche[1] + marge)
    ax1.set_title("Zoom - Coin Supérieur Gauche", fontsize=14, fontweight='bold')
    ax1.set_xlabel("X")
    ax1.set_ylabel("Y")
    
    # Zoom adaptatif sur le coin supérieur droit
    quad4 = quadrilateres[2]
    x4, y4 = quad4.exterior.xy
    
    # Trouver le coin supérieur droit (point avec x max et y max)
    points4 = list(zip(x4, y4))
    coin_haut_droit = builtins.max(points4, key=lambda p: p[0] + p[1])  # x maximum, y maximum
    
    # Calculer la taille de la zone de zoom basée sur la taille du quadrilatère
    largeur_quad4 = builtins.max(x4) - builtins.min(x4)
    hauteur_quad4 = builtins.max(y4) - builtins.min(y4)
    marge = builtins.max(largeur_quad4, hauteur_quad4) * 0.01  # 1% de marge
    
    ax2.set_xlim(coin_haut_droit[0] - marge, coin_haut_droit[0] + marge)
    ax2.set_ylim(coin_haut_droit[1] - marge, coin_haut_droit[1] + marge)
    ax2.set_title("Zoom - Coin Supérieur Droit", fontsize=14, fontweight='bold')
    ax2.set_xlabel("X")
    ax2.set_ylabel("Y")
    
    plt.tight_layout()
    return fig

#----------------------Edition du rapport pdf FORMAT A4 PAYSAGE---------------

def creer_page_complete(ligne_data, quadrilateres):
    """Crée une page A4 PAYSAGE avec les détails et les deux graphiques côte à côte"""
    # Format A4 en PAYSAGE : 11.69 x 8.27 pouces
    fig = plt.figure(figsize=(11.69, 8.27))
    
    # Titre de la page
    fig.suptitle(f"Vitrage {ligne_data['ID vitrage']}", 
                 fontsize=14, fontweight='bold', y=0.95)
    
    # Créer une grille pour organiser le contenu en PAYSAGE
    # 3 lignes, 2 colonnes : [Métriques sur toute la largeur] [Graph gauche | Graph droit]
    gs = fig.add_gridspec(3, 2, height_ratios=[0.25, 0.6, 0.15],hspace=0.3, wspace=0.20,
                          left=0.06, right=0.97, top=0.90, bottom=0.06)
    
    # Section 1: Tableau des métriques principales (en haut, sur toute la largeur)
    ax_metrics = fig.add_subplot(gs[0, :])
    ax_metrics.axis('tight')
    ax_metrics.axis('off')

   # Récupérer les valeurs des écarts
    ecart_gauche = ligne_data['Ecart minimal // bornes gauche (mm)']
    ecart_droite = ligne_data['Ecart minimal // bornes droite (mm)']
    ecart_hautes = ligne_data['Ecart minimal // bornes hautes (mm)']

    metrics_data = [
        ['Paramètre', 'Valeur', 'Unité'],
        ['Gamme', f"{ligne_data['Gamme']}", '-'],
        ['Demi-périmètre', f"{ligne_data['Demi-périmètre (m)']:.1f}", 'm'],
        ['Ecart minimal // bornes gauche', f"{ligne_data['Ecart minimal // bornes gauche (mm)']:.1f}", 'mm'],
        ['Ecart minimal // bornes droite', f"{ligne_data['Ecart minimal // bornes droite (mm)']:.1f}", 'mm'],
        ['Ecart minimal // bornes hautes', f"{ligne_data['Ecart minimal // bornes hautes (mm)']:.1f}", 'mm']
    ]
    
    table = ax_metrics.table(cellText=metrics_data, cellLoc='center', loc='center',
                            colWidths=[0.3, 0.1, 0.1])
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.scale(1, 1.5)
    
    # Style du tableau
    # Style du tableau avec code couleur pour les écarts
    for i in range(len(metrics_data)):
        for j in range(3):
            cell = table[(i, j)]
            if i == 0:  # En-tête
                cell.set_facecolor("#006D8F")
                cell.set_text_props(weight='bold', color='white')
            elif i == 3 and j == 1:  # Ecart gauche
                if ecart_gauche < 0:
                    cell.set_facecolor('#FFB3B3')  # Rouge clair
                else:
                    cell.set_facecolor('#B3FFB3')  # Vert clair
            elif i == 4 and j == 1:  # Ecart droite
                if ecart_droite < 0:
                    cell.set_facecolor('#FFB3B3')  # Rouge clair
                else:
                    cell.set_facecolor('#B3FFB3')  # Vert clair
            elif i == 5 and j == 1:  # Ecart hautes
                if ecart_hautes < 0:
                    cell.set_facecolor('#FFB3B3')  # Rouge clair
                else:
                    cell.set_facecolor('#B3FFB3')  # Vert clair
            else:
                cell.set_facecolor('#E7E6E6' if i == 1 or i == 2 else 'white')

    couleurs = ["#000000","#FF3333", '#006D8F', "#FF3333"]
    titre_graph = ['Cadre','Limite','Vitrage','Limite']
    
    # Section 3: GRAPHIQUE GAUCHE - Zoom coin supérieur gauche
    ax1 = fig.add_subplot(gs[1, 0])
    
    for i, quad in enumerate(quadrilateres):
        x, y = quad.exterior.xy
        patch = MplPolygon(list(zip(x, y)), facecolor='none', 
                         edgecolor=couleurs[i], linewidth=1.5,
                         label=titre_graph[i])
        ax1.add_patch(patch)
    
    ax1.set_aspect('equal')
    ax1.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
    ax1.legend(loc='upper right', fontsize=8, framealpha=0.9)
    
    # Zoom adaptatif sur le coin supérieur gauche
    quad1 = quadrilateres[2]
    x1, y1 = quad1.exterior.xy
    points1 = list(zip(x1, y1))
    coin_haut_gauche = builtins.min(points1, key=lambda p: p[0] - p[1])
    
    largeur_quad1 = builtins.max(x1) - builtins.min(x1)
    hauteur_quad1 = builtins.max(y1) - builtins.min(y1)
    marge = builtins.max(largeur_quad1, hauteur_quad1) * 0.01
    
    ax1.set_xlim(coin_haut_gauche[0] - marge, coin_haut_gauche[0] + marge)
    ax1.set_ylim(coin_haut_gauche[1] - marge, coin_haut_gauche[1] + marge)
    ax1.set_title("Coin Supérieur Gauche", 
                  fontsize=11, fontweight='bold', pad=10)
    ax1.set_xlabel("Coordonnée X (mm)", fontsize=9)
    ax1.set_ylabel("Coordonnée Y (mm)", fontsize=9)
    ax1.tick_params(labelsize=8)
    
    # Section 4: GRAPHIQUE DROIT - Zoom coin supérieur droit
    ax2 = fig.add_subplot(gs[1, 1])
    
    for i, quad in enumerate(quadrilateres):
        x, y = quad.exterior.xy
        patch = MplPolygon(list(zip(x, y)), facecolor='none', 
                         edgecolor=couleurs[i], linewidth=1.5,
                         label=titre_graph[i])
        ax2.add_patch(patch)
    
    ax2.set_aspect('equal')
    ax2.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
    ax2.legend(loc='upper right', fontsize=8, framealpha=0.9)
    
    # Zoom adaptatif sur le coin supérieur droit
    quad4 = quadrilateres[2]
    x4, y4 = quad4.exterior.xy
    points4 = list(zip(x4, y4))
    coin_haut_droit = builtins.max(points4, key=lambda p: p[0] + p[1])
    
    largeur_quad4 = builtins.max(x4) - builtins.min(x4)
    hauteur_quad4 = builtins.max(y4) - builtins.min(y4)
    marge = builtins.max(largeur_quad4, hauteur_quad4) * 0.01
    
    ax2.set_xlim(coin_haut_droit[0] - marge, coin_haut_droit[0] + marge)
    ax2.set_ylim(coin_haut_droit[1] - marge, coin_haut_droit[1] + marge)
    ax2.set_title("Coin Supérieur Droit", 
                  fontsize=11, fontweight='bold', pad=10)
    ax2.set_xlabel("Coordonnée X (mm)", fontsize=9)
    ax2.set_ylabel("Coordonnée Y (mm)", fontsize=9)
    ax2.tick_params(labelsize=8)

    # Section 2: Annotation (sur toute la largeur)
    ax_annotation = fig.add_subplot(gs[2, :])
    ax_annotation.axis('off')
    
    cadre = quadrilateres[0]
    A_c, B_c, C_c, D_c, origine_c = cadre.exterior.coords
    
    if D_c[1] > C_c[1]:
        Sens_y = 'droite'
        opp = 'gauche'
    else:
        Sens_y = 'gauche'
        opp = 'doite'
    if A_c[0] > D_c[0]:
        Sens_x = 'droite'
    else:
        Sens_x = 'gauche'

    if round(builtins.max(abs(A_c[0]-D_c[0]),abs(B_c[0]-C_c[0])),1) == 0:
        annotation_text = f"""Le montant de {Sens_y} descend de {round(abs(D_c[1]-C_c[1]),1)}mm par rapport au montant de {opp}.\n"""
    elif round(abs(D_c[1]-C_c[1]),1) == 0:
        annotation_text = f""" La traverse haute se déplace de {round(builtins.max(abs(A_c[0]-D_c[0]),abs(B_c[0]-C_c[0])),1)}mm vers la {Sens_x} par rapport à la traverse basse."""
    else:
        annotation_text = f"""Le montant de {Sens_y} descend de {round(abs(D_c[1]-C_c[1]),1)}mm par rapport au montant de {opp}.\n La traverse haute se déplace de {round(builtins.max(abs(A_c[0]-D_c[0]),abs(B_c[0]-C_c[0])),1)}mm vers la {Sens_x} par rapport à la traverse basse."""

    ax_annotation.text(0.5, 0.5, annotation_text, 
                      ha='center', va='center', 
                      fontsize=9, family='Sans',
                      bbox=dict(boxstyle='square,pad=0.3', facecolor="#EDF4F5", 
                               edgecolor="#006D8F", linewidth=1, alpha=0.9))
     
    return fig

# Fonction pour générer le rapport PDF complet
def generer_rapport_pdf(df, graphs):
    """Génère un rapport PDF avec toutes les lignes au format A4 PAYSAGE"""
    buffer = io.BytesIO()
    
    with PdfPages(buffer) as pdf:
        # Page de titre - Format A4 PAYSAGE
        fig_titre = plt.figure(figsize=(11.69, 8.27))
        
        # En-tête institutionnel
        fig_titre.text(0.5, 0.65, "ANALYSE DE LA MISE EN PARALLELOGRAMME\n DES VITRAGES", 
                       ha='center', va='center', 
                       fontsize=16, fontweight='bold', family='Sans')
        
        # Ligne de séparation
        fig_titre.text(0.5, 0.60, "─" * 80, 
                       ha='center', va='center', 
                       fontsize=10)

        #Mettre une image
        url_logo = 'https://github.com/lilianmtech/Analyse_Parallelogramme_Vitrage/blob/main/logo-couleur.png?raw=true'
        response = requests.get(url_logo)
        img = mpimg.imread(io.BytesIO(response.content))
        orig_width, orig_height = img.getSize()
        scale = 0.5
        img_width = orig_width * scale
        img_height = orig_height * scale
        ax = fig_titre.add_axes([0.38, 0.43, 0.25, 0.1375])  # [left, bottom, img_width, img_height]
        ax.imshow(img)
        ax.axis("off")

        # Résumer
        mask = (df[["Ecart minimal // bornes gauche (mm)", "Ecart minimal // bornes droite (mm)", "Ecart minimal // bornes hautes (mm)"]] < 0).any(axis=1)
        nb_non_conforme = mask.sum()
        nb_conforme = len(df)-nb_non_conforme
        
        Gammes = df["Gamme"].value_counts().index
        i=0
        for g in Gammes:
            if i==0:
                Gamme = str(g)
            else :
                Gamme += f" & {g}"
            i = i+1
        
        print(Gammes)
        stats_text = f"""
Gamme Raico de l'étude: {Gamme}
Nombre vitrage étudiés: {len(df)}
Nombre de vitrage conforme: {nb_conforme}
Nombre de vitrage non-conforme: {nb_non_conforme}
        """
        
        fig_titre.text(0.5, 0.35, stats_text, 
                       ha='center', va='center', 
                       fontsize=10, family='Sans', style='normal',
                       bbox=dict(boxstyle='round', facecolor='#E8F4F8', alpha=0.8))
        
        # Pied de page
        from datetime import datetime
        date_rapport = datetime.now().strftime("%d/%m/%Y")
        fig_titre.text(0.5, 0.25, f"Date du rapport: {date_rapport}", 
                       ha='center', va='center', 
                       fontsize=11, family='Sans')

        fig_titre.text(0.5, 0.15, "MTECHBUILD", 
                       ha='center', va='center', 
                       fontsize=10, style='italic', color='gray', family='Sans')
        
        plt.axis('off')
        pdf.savefig(fig_titre, bbox_inches='tight')
        plt.close(fig_titre)
        
        # Pour chaque ligne - une page A4 PAYSAGE avec les deux graphiques côte à côte
        for idx, row in df.iterrows():
            quadrilateres = graphs["Graph"][idx]
            fig_complete = creer_page_complete(row, quadrilateres)
            pdf.savefig(fig_complete, bbox_inches='tight')
            plt.close(fig_complete)
        
        # Métadonnées du PDF
        d = pdf.infodict()
        d['Title'] = 'Analyse de la mise en paraléllogramme des vitrages'
        d['Author'] = 'MTECHBUILD'
        d['Subject'] = 'Analyse de la mise en paraléllogramme des vitrages'
        d['Keywords'] = 'Vitrage, Analyse, Paraléllogramme'
        d['Creator'] = 'Application Streamlit'
    
    buffer.seek(0)
    return buffer


def Ajout_Titre(input_pdf, watermark_url, transparency, scale, pos_y, pos_x):
    reader = PdfReader(input_pdf)
    writer = PdfWriter()

    # Télécharger l'image depuis GitHub (raw URL)
    response = requests.get(watermark_url)
    img_data = io.BytesIO(response.content)
    img = ImageReader(img_data)

    # Dimensions originales de l'image
    orig_width, orig_height = img.getSize()

    for i, page in enumerate(reader.pages):
        # Ne pas appliquer sur la première page
        if i == 0:
            writer.add_page(page)
            continue

        largeur = float(page.mediabox.width)
        hauteur = float(page.mediabox.height)

        packet = io.BytesIO()
        c = canvas.Canvas(packet, pagesize=(largeur, hauteur))

        x = pos_x * largeur
        y = pos_y * hauteur

        c.setFillAlpha(transparency)

        # Conserver proportions : multiplier largeur et hauteur originales par le facteur scale
        img_width = orig_width * scale
        img_height = orig_height * scale

        # Centrer l'image autour de (x, y)
        c.drawImage(img, x - img_width/2, y - img_height/2,
                    width=img_width, height=img_height, mask='auto')

        c.save()

        packet.seek(0)
        watermark = PdfReader(packet)
        page.merge_page(watermark.pages[0])
        writer.add_page(page)

    output = io.BytesIO()
    writer.write(output)
    output.seek(0)
    return output

def colorer_valeurs(val):
    if val > 0:
        color = "green"
    elif val < 0:
        color = "red"
    else:
        color = "orange"
    return f"color: {color}; font-weight: bold;"
#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

# Configuration de la page
st.set_page_config(page_title="Analyse de la mise en parallélogramme", layout="wide")

# Titre de l'application
st.title("🏢 Analyse de la mise en parallélogramme des vitrages")

choix = st.selectbox("Calage latéral :", ["Sans", "Avec"])

# --- Ligne 1 ---
ligne1_col1, ligne1_col2 = st.columns(2)
url = 'https://github.com/lilianmtech/Analyse_Parallelogramme_Vitrage/blob/main/'
with ligne1_col1:
    if choix == "Sans":
        st.image(url+"Borne_M.png"+'?raw=true', caption="Bornes sur montants", width=450)
    elif choix == "Avec":
        st.image(url+"Borne_M_C.png"+'?raw=true', caption="Bornes sur montants", width=450)


with ligne1_col2:
    if choix == "Sans":
        st.image(url+"Borne_T.png"+'?raw=true', caption="Bornes sur traverses", width=400)
    elif choix == "Avec":
        st.image(url+"Borne_T_C.png"+'?raw=true', caption="Bornes sur traverses", width=400)   

# --- Ligne 2 ---
ligne2_col1, ligne2_col2 = st.columns(2)

with ligne2_col1:
    if choix == "Sans":
        Tm1 = st.number_input("Valeur Tolérance Tm1 :", value=2.5, step=0.1, format="%.2f")
        Tm2 = st.number_input("Valeur Tolérance Tm2 + menuiserie :", value=10.5, step=0.1, format="%.2f")
    else:
        Tm =  st.number_input("Valeur Tolérance Tm :", value=2.5, step=0.1, format="%.2f")
        Ca =  st.number_input("Epaisseur cale C + menuiserie :", value=13.0, step=0.1, format="%.2f")
        Jc =  st.number_input("Jeu entre vitrage et cale Jc :", value=2.0, step=0.1, format="%.2f")
    
with ligne2_col2:
    if choix == "Sans":
        Tt1 = st.number_input("Valeur Tolérance Tt1 :", value=5.0, step=0.1, format="%.2f")
        Tt2 = st.number_input("Valeur Tolérance Tt2 + menuiserie :", value=13.0, step=0.1, format="%.2f")
    else:
        Tt1 =  st.number_input("Valeur Tolérance Tt1 :", value=5.0, step=0.1, format="%.2f")
        Tt2 =  st.number_input("Valeur Tolérance Tt2 + menuiserie :", value=13.0, step=0.1, format="%.2f")
        Jv =  st.number_input("Jeu entre borne basse et vitrage Jv :", value=1.0, step=0.1, format="%.2f")


#---------------------------Importation des données---------------------------

st.header("📂 Importer Fichier Excel")

uploaded_file = st.file_uploader("Importer un fichier de données (CSV ou Excel)", type=["csv", "xlsx"])

#---------------------------Chargement des vitrages---------------------------
if choix == "Sans":
    Raico = [Tm1, Tm2, Tt1, Tt2]

else:
    Raico = [Tm, Ca, Jc, Tt1, Tt2, Jv]


if uploaded_file:
    # Mémoriser le nom du fichier pour savoir si un nouveau fichier est importé
    if "uploaded_name" not in st.session_state or st.session_state["uploaded_name"] != uploaded_file.name:
        st.session_state.clear()
        st.session_state["uploaded_name"] = uploaded_file.name

    # Vérifier si un changement de paramètre nécessite une mise à jour
    params_actuels = {
        "choix": choix,
        "Raico": Raico,
    }

    if (
        "params_prec" not in st.session_state
        or st.session_state["params_prec"] != params_actuels
    ):
        datas, graphs = Dechaussement_vitrage(uploaded_file, choix, Raico)
        st.session_state["datas"] = datas
        st.session_state["graphs"] = graphs
        st.session_state["params_prec"] = params_actuels
    else:
        datas = st.session_state["datas"]
        graphs = st.session_state["graphs"]

    # -------------------- Affichage du tableau --------------------
    st.header("📊 Tableau des Résultats")
    df_affichage = pd.DataFrame(datas)
    
    styled_df = (
    df_affichage.style
    .map(colorer_valeurs, subset=["Ecart minimal // bornes gauche (mm)", 
                                  "Ecart minimal // bornes droite (mm)", 
                                  "Ecart minimal // bornes hautes (mm)"])
    .format(precision=1)
    .set_properties(**{'text-align': 'center'})
    .set_properties(**{'text-align': 'center', 'vertical-align': 'middle'})
    .set_table_styles([
        {"selector": "th", "props": [("text-align", "center")]},
        {"selector": "td", "props": [("text-align", "center")]}
    ])
)
    st.dataframe(styled_df, width='stretch', hide_index=True)

    # -------------------- Sélection et visualisation --------------------
    st.header("🔲 Sélection et Visualisation")
    col1, col2 = st.columns([1, 3])

    with col1:
        ligne_selectionnee = st.selectbox(
            "Sélectionnez le vitrage à visualiser :",
            options=list(datas["ID vitrage"]),
            key="select_vitrage",
        )

        idx = datas["ID vitrage"].index(ligne_selectionnee)
        ligne_data = {col: datas[col][idx] for col in datas.keys()}

        for col in ligne_data.keys():
                if col != 'ID vitrage':
                    st.metric(col, ligne_data[col])

    with col2:
        idx = graphs["ID vitrage"].index(ligne_selectionnee)
        quadrilateres = graphs["Graph"][idx]
        fig = visualiser_quadrilateres(quadrilateres)
        st.pyplot(fig)
        # Informations supplémentaires
        cadre = graphs["Graph"][idx][0]
        A_c,B_c,C_c,D_c,origine_c = cadre.exterior.coords
        if D_c[1]>C_c[1]:
            Sens_y = 'droite'
        else:
            Sens_y = 'gauche'
        if A_c[0]>D_c[0]:
            Sens_x = 'droite'
        else:
            Sens_x = 'gauche'
        

        cadre = quadrilateres[0]
    A_c, B_c, C_c, D_c, origine_c = cadre.exterior.coords
    
    if D_c[1] > C_c[1]:
        Sens_y = 'droite'
        opp = 'gauche'
    else:
        Sens_y = 'gauche'
        opp = 'droite'
    if A_c[0] > D_c[0]:
        Sens_x = 'droite'
    else:
        Sens_x = 'gauche'

    if round(builtins.max(abs(A_c[0]-D_c[0]),abs(B_c[0]-C_c[0])),1) == 0:
        st.info(f"""📐
        Le montant de **{Sens_y}** descend de **{round(abs(D_c[1]-C_c[1]),1)}mm** par rapport au montant de {opp}.
        """)
    elif round(abs(D_c[1]-C_c[1]),1) == 0:
        st.info(f"""📐
        La traverse haute se déplace de **{round(builtins.max(abs(A_c[0]-D_c[0]),abs(B_c[0]-C_c[0])),1)}mm** vers la **{Sens_x}** par rapport à la traverse basse.
        """)
    else:
        st.info(f"""📐
        Le montant de **{Sens_y}** descend de **{round(abs(D_c[1]-C_c[1]),1)}mm** par rapport au montant de {opp}.\n La traverse haute se déplace de **{round(builtins.max(abs(A_c[0]-D_c[0]),abs(B_c[0]-C_c[0])),1)}mm** vers la **{Sens_x}** par rapport à la traverse basse.
        """)

    
    st.divider()
    col_pdf1, col_pdf2, col_pdf3 = st.columns([1, 2, 1])
    with col_pdf2:
        st.subheader("📄 Rapport PDF Complet")
        st.write("Téléchargez un rapport PDF contenant les détails et visualisations de toutes les lignes")
        
        # Bouton pour générer le PDF
        if st.button("🔄 Générer le Rapport PDF", use_container_width=True, type="primary"):
            with st.spinner("Génération du rapport PDF en cours..."):
                try:
                    pdf = generer_rapport_pdf(df_affichage, graphs)
                    pdf_buffer=Ajout_Titre(pdf, 'https://github.com/lilianmtech/Analyse_Parallelogramme_Vitrage/blob/main/logo-couleur.png?raw=true', 0.3, 0.3, 0.9, 0.9)
                    st.session_state['pdf_buffer'] = pdf_buffer
                    st.session_state['pdf_generated'] = True
                    st.success("✅ Rapport PDF généré avec succès !")
                except Exception as e:
                    st.error(f"❌ Erreur lors de la génération : {str(e)}")
        
        # Bouton de téléchargement si le PDF a été généré
        if st.session_state.get('pdf_generated', False):
            st.download_button(
                label="⬇️ Télécharger le Rapport PDF",
                data=st.session_state['pdf_buffer'],
                file_name="rapport_quadrilateres.pdf",
                mime="application/pdf",
                use_container_width=True,
                key="download_pdf_button"
            )
            st.info(f"📋 Le rapport contient {len(df_affichage) + 1} pages")

else:
    st.info("📥 Importez un fichier Excel pour commencer l’analyse.")
        # Footer
st.caption("Application développée avec Streamlit et Shapely")





























